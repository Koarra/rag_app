import json
import csv
import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def json_to_excel_with_table(json_path, excel_path):
    # Ensure input file exists
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"JSON file not found: {json_path}")

    # Load JSON data
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not data:
        raise ValueError("JSON file is empty.")

    # Determine headers (all keys except the "person" key)
    inner_keys = list(next(iter(data.values())).keys())
    fieldnames = ["person"] + inner_keys

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write headers
    ws.append(fieldnames)

    # Write data rows
    for person, info in data.items():
        row = [person] + [info.get(k, "") for k in inner_keys]
        ws.append(row)

    # Define table range
    num_rows = ws.max_row
    num_cols = ws.max_column

    # Handle Excel-style column naming beyond Z
    def col_letter(n):
        result = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            result = chr(65 + remainder) + result
        return result

    table_range = f"A1:{col_letter(num_cols)}{num_rows}"

    # Create table
    table = Table(displayName="MyTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save Excel
    wb.save(excel_path)
    print(f"Excel file saved at: {excel_path}")
